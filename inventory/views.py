from datetime import timedelta
from math import ceil

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import F, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font

from .forms import ProductForm, SaleForm, StockReceiptForm
from .models import ActionLog, Category, Product, Sale, StockReceipt


def _forbidden(request, message='У вас нет прав для доступа к данной странице.'):
    return render(request, 'inventory/403.html', {'message': message}, status=403)


def _is_admin(user):
    return user.is_superuser or user.groups.filter(name='Администратор').exists()


def _is_merchandiser(user):
    return user.groups.filter(name='Мерчендайзер').exists()


def _is_seller(user):
    return user.groups.filter(name='Продавец').exists()


def _log_action(user, action_type, product=None, description=""):
    ActionLog.objects.create(
        user=user if user.is_authenticated else None,
        action_type=action_type,
        product=product,
        description=description
    )


def _make_excel_response(workbook, filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def _style_header(sheet, row_number=1):
    for cell in sheet[row_number]:
        cell.font = Font(bold=True)


def _autosize_columns(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = str(cell.value) if cell.value is not None else ""
            if len(value) > max_length:
                max_length = len(value)
        sheet.column_dimensions[column_letter].width = max_length + 2


def _calculate_replenishment_metrics(product):
    date_from = timezone.now() - timedelta(days=30)

    sold_last_30_days = (
        Sale.objects.filter(product=product, sale_date__gte=date_from)
        .aggregate(total=Coalesce(Sum('quantity'), 0))['total']
    )

    avg_daily_sales = sold_last_30_days / 30 if sold_last_30_days else 0

    if avg_daily_sales > 0:
        days_left = round(product.quantity / avg_daily_sales, 1)
    else:
        days_left = None

    base_replenishment = max(product.minimum_quantity - product.quantity, 0)

    target_stock = product.minimum_quantity
    if avg_daily_sales > 0:
        target_stock = max(product.minimum_quantity, ceil(avg_daily_sales * 14))

    recommended_order = max(target_stock - product.quantity, base_replenishment)

    if days_left is None:
        priority = "Низкий"
    elif days_left <= 3:
        priority = "Высокий"
    elif days_left <= 7:
        priority = "Средний"
    else:
        priority = "Низкий"

    product.sold_last_30_days = sold_last_30_days
    product.avg_daily_sales = round(avg_daily_sales, 2)
    product.days_left = days_left
    product.recommended_order = recommended_order
    product.replenishment_priority = priority

    return product


@login_required
def role_redirect(request):
    return redirect('home')


@login_required
def home(request):
    if not (_is_admin(request.user) or _is_merchandiser(request.user) or _is_seller(request.user)):
        return _forbidden(request)

    products_count = Product.objects.count()
    categories_count = Category.objects.count()
    low_stock_count = Product.objects.filter(quantity__lte=F('minimum_quantity')).count()

    top_product = (
        Sale.objects.values('product__name')
        .annotate(total_sold=Coalesce(Sum('quantity'), 0))
        .order_by('-total_sold')
        .first()
    )

    products_for_chart = Product.objects.all().order_by('name')
    chart_labels = [product.name for product in products_for_chart]
    chart_data = [product.quantity for product in products_for_chart]

    # Товары без движения за 30 дней
    no_movement_date_from = timezone.now() - timedelta(days=30)
    sold_product_ids = Sale.objects.filter(
        sale_date__gte=no_movement_date_from
    ).values_list('product_id', flat=True)

    no_movement_products = Product.objects.select_related('category').exclude(
        id__in=sold_product_ids
    ).order_by('name')[:5]

    no_movement_count = Product.objects.exclude(id__in=sold_product_ids).count()

    # Последние продажи
    latest_sales = Sale.objects.select_related('product').order_by('-sale_date')[:5]

    # Последние поступления
    latest_receipts = StockReceipt.objects.select_related('product').order_by('-receipt_date')[:5]

    # Товары с высоким приоритетом пополнения
    replenishment_queryset = Product.objects.select_related('category', 'supplier').filter(
        quantity__lt=F('minimum_quantity')
    ).order_by('quantity', 'name')

    replenishment_products = []
    for product in replenishment_queryset:
        replenishment_products.append(_calculate_replenishment_metrics(product))

    priority_order = {'Высокий': 0, 'Средний': 1, 'Низкий': 2}
    replenishment_products.sort(
        key=lambda item: (priority_order.get(item.replenishment_priority, 3), item.quantity, item.name)
    )
    top_replenishment_products = replenishment_products[:5]

    context = {
        'products_count': products_count,
        'categories_count': categories_count,
        'low_stock_count': low_stock_count,
        'no_movement_count': no_movement_count,
        'top_product': top_product,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'top_replenishment_products': top_replenishment_products,
        'no_movement_products': no_movement_products,
        'latest_sales': latest_sales,
        'latest_receipts': latest_receipts,
    }
    return render(request, 'inventory/home.html', context)


@login_required
def product_list(request):
    if not (_is_admin(request.user) or _is_seller(request.user) or _is_merchandiser(request.user)):
        return _forbidden(request)

    query = request.GET.get('q', '').strip()
    category_id = request.GET.get('category', '').strip()

    categories_queryset = Category.objects.all().order_by('name')
    products_queryset = Product.objects.select_related('category', 'supplier').all().order_by('name')

    if category_id:
        products_queryset = products_queryset.filter(category_id=category_id)

    products = list(products_queryset)

    if query:
        query_lower = query.lower()
        filtered_products = []

        for product in products:
            name_value = (product.name or '').lower()
            article_value = (product.article or '').lower()

            if query_lower in name_value or query_lower in article_value:
                filtered_products.append(product)

        products = filtered_products

    categories = []
    for category in categories_queryset:
        categories.append({
            'id': category.id,
            'name': category.name,
            'selected': str(category.id) == category_id
        })

    context = {
        'products': products,
        'categories': categories,
        'query': query,
    }
    return render(request, 'inventory/product_list.html', context)


@login_required
def export_products_excel(request):
    if not (_is_admin(request.user) or _is_merchandiser(request.user) or _is_seller(request.user)):
        return _forbidden(request)

    query = request.GET.get('q', '').strip()
    category_id = request.GET.get('category', '').strip()

    products_queryset = Product.objects.select_related('category', 'supplier').all().order_by('name')

    if category_id:
        products_queryset = products_queryset.filter(category_id=category_id)

    products = list(products_queryset)

    if query:
        query_lower = query.lower()
        filtered_products = []

        for product in products:
            name_value = (product.name or '').lower()
            article_value = (product.article or '').lower()

            if query_lower in name_value or query_lower in article_value:
                filtered_products.append(product)

        products = filtered_products

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Товары'

    sheet.append([
        'Наименование',
        'Категория',
        'Поставщик',
        'Пол',
        'Размер',
        'Цвет',
        'Артикул',
        'Цена',
        'Количество',
        'Минимальный остаток',
        'Статус'
    ])

    for product in products:
        sheet.append([
            product.name,
            str(product.category),
            str(product.supplier) if product.supplier else '',
            product.gender,
            product.size,
            product.color,
            product.article,
            float(product.price),
            product.quantity,
            product.minimum_quantity,
            product.stock_status
        ])

    _style_header(sheet)
    _autosize_columns(sheet)

    return _make_excel_response(workbook, 'products.xlsx')


@login_required
def create_product(request):
    if not _is_admin(request.user):
        return _forbidden(request, 'Добавление товара доступно только администратору.')

    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save()

            _log_action(
                user=request.user,
                action_type="CREATE_PRODUCT",
                product=product,
                description=f'Добавлен товар "{product.name}"'
            )

            messages.success(request, 'Товар успешно добавлен.')
            return redirect('product_detail', product_id=product.id)
    else:
        form = ProductForm()

    context = {
        'form': form,
    }
    return render(request, 'inventory/create_product.html', context)


@login_required
def product_detail(request, product_id):
    if not (_is_admin(request.user) or _is_merchandiser(request.user) or _is_seller(request.user)):
        return _forbidden(request)

    product = get_object_or_404(Product.objects.select_related('category', 'supplier'), id=product_id)

    receipts_count = StockReceipt.objects.filter(product=product).count()
    sales_count = Sale.objects.filter(product=product).count()

    context = {
        'product': product,
        'receipts_count': receipts_count,
        'sales_count': sales_count,
    }
    return render(request, 'inventory/product_detail.html', context)


@login_required
def edit_product(request, product_id):
    if not _is_admin(request.user):
        return _forbidden(request, 'Редактирование товара доступно только администратору.')

    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            updated_product = form.save()

            _log_action(
                user=request.user,
                action_type="UPDATE_PRODUCT",
                product=updated_product,
                description=f'Изменён товар "{updated_product.name}"'
            )

            messages.success(request, 'Товар успешно обновлён.')
            return redirect('product_detail', product_id=product.id)
    else:
        form = ProductForm(instance=product)

    context = {
        'form': form,
        'product': product,
    }
    return render(request, 'inventory/edit_product.html', context)


@login_required
def delete_product(request, product_id):
    if not _is_admin(request.user):
        return _forbidden(request, 'Удаление товара доступно только администратору.')

    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        product_name = product.name

        _log_action(
            user=request.user,
            action_type="DELETE_PRODUCT",
            product=None,
            description=f'Удалён товар "{product_name}"'
        )

        product.delete()
        messages.success(request, f'Товар "{product_name}" успешно удалён.')
        return redirect('product_list')

    context = {
        'product': product,
    }
    return render(request, 'inventory/delete_product.html', context)


@login_required
def stock_list(request):
    if not (_is_admin(request.user) or _is_merchandiser(request.user) or _is_seller(request.user)):
        return _forbidden(request)

    products = Product.objects.select_related('category', 'supplier').all().order_by('name')
    return render(request, 'inventory/stock_list.html', {'products': products})


@login_required
def low_stock(request):
    if not (_is_admin(request.user) or _is_merchandiser(request.user)):
        return _forbidden(request)

    products = Product.objects.select_related('category', 'supplier').filter(
        quantity__lte=F('minimum_quantity')
    ).order_by('quantity', 'name')

    return render(request, 'inventory/low_stock.html', {'products': products})


@login_required
def export_low_stock_excel(request):
    if not (_is_admin(request.user) or _is_merchandiser(request.user)):
        return _forbidden(request)

    products = Product.objects.select_related('category', 'supplier').filter(
        quantity__lte=F('minimum_quantity')
    ).order_by('quantity', 'name')

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Низкий остаток'

    sheet.append([
        'Наименование',
        'Категория',
        'Поставщик',
        'Пол',
        'Размер',
        'Цвет',
        'Артикул',
        'Количество',
        'Минимальный остаток',
        'Статус'
    ])

    for product in products:
        sheet.append([
            product.name,
            str(product.category),
            str(product.supplier) if product.supplier else '',
            product.gender,
            product.size,
            product.color,
            product.article,
            product.quantity,
            product.minimum_quantity,
            product.stock_status
        ])

    _style_header(sheet)
    _autosize_columns(sheet)

    return _make_excel_response(workbook, 'low_stock.xlsx')


@login_required
def replenishment_list(request):
    if not (_is_admin(request.user) or _is_merchandiser(request.user)):
        return _forbidden(request)

    products_queryset = Product.objects.select_related('category', 'supplier').filter(
        quantity__lt=F('minimum_quantity')
    ).order_by('quantity', 'name')

    products = []
    for product in products_queryset:
        products.append(_calculate_replenishment_metrics(product))

    priority_order = {'Высокий': 0, 'Средний': 1, 'Низкий': 2}
    products.sort(key=lambda item: (priority_order.get(item.replenishment_priority, 3), item.quantity, item.name))

    return render(request, 'inventory/replenishment_list.html', {'products': products})


@login_required
def export_replenishment_excel(request):
    if not (_is_admin(request.user) or _is_merchandiser(request.user)):
        return _forbidden(request)

    products_queryset = Product.objects.select_related('category', 'supplier').filter(
        quantity__lt=F('minimum_quantity')
    ).order_by('quantity', 'name')

    products = []
    for product in products_queryset:
        products.append(_calculate_replenishment_metrics(product))

    priority_order = {'Высокий': 0, 'Средний': 1, 'Низкий': 2}
    products.sort(key=lambda item: (priority_order.get(item.replenishment_priority, 3), item.quantity, item.name))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Пополнение'

    sheet.append([
        'Наименование',
        'Категория',
        'Поставщик',
        'Пол',
        'Размер',
        'Цвет',
        'Артикул',
        'Текущий остаток',
        'Минимальный остаток',
        'Средние продажи в день',
        'Хватит на дней',
        'Рекомендуемый заказ',
        'Приоритет'
    ])

    for product in products:
        sheet.append([
            product.name,
            str(product.category),
            str(product.supplier) if product.supplier else '',
            product.gender,
            product.size,
            product.color,
            product.article,
            product.quantity,
            product.minimum_quantity,
            product.avg_daily_sales,
            product.days_left if product.days_left is not None else 'Нет продаж',
            product.recommended_order,
            product.replenishment_priority
        ])

    _style_header(sheet)
    _autosize_columns(sheet)

    return _make_excel_response(workbook, 'replenishment.xlsx')


@login_required
def no_movement_products(request):
    if not (_is_admin(request.user) or _is_merchandiser(request.user)):
        return _forbidden(request)

    period_days = request.GET.get('days', '30').strip()

    try:
        period_days_int = int(period_days)
    except ValueError:
        period_days_int = 30

    if period_days_int not in [30, 60, 90]:
        period_days_int = 30

    date_from = timezone.now() - timedelta(days=period_days_int)

    sold_product_ids = Sale.objects.filter(
        sale_date__gte=date_from
    ).values_list('product_id', flat=True)

    products = Product.objects.select_related('category', 'supplier').exclude(
        id__in=sold_product_ids
    ).order_by('name')

    period_options = [
        {'value': 30, 'label': 'За 30 дней', 'selected': period_days_int == 30},
        {'value': 60, 'label': 'За 60 дней', 'selected': period_days_int == 60},
        {'value': 90, 'label': 'За 90 дней', 'selected': period_days_int == 90},
    ]

    context = {
        'products': products,
        'period_days': period_days_int,
        'period_options': period_options,
    }
    return render(request, 'inventory/no_movement_products.html', context)


@login_required
def export_no_movement_excel(request):
    if not (_is_admin(request.user) or _is_merchandiser(request.user)):
        return _forbidden(request)

    period_days = request.GET.get('days', '30').strip()

    try:
        period_days_int = int(period_days)
    except ValueError:
        period_days_int = 30

    if period_days_int not in [30, 60, 90]:
        period_days_int = 30

    date_from = timezone.now() - timedelta(days=period_days_int)

    sold_product_ids = Sale.objects.filter(
        sale_date__gte=date_from
    ).values_list('product_id', flat=True)

    products = Product.objects.select_related('category', 'supplier').exclude(
        id__in=sold_product_ids
    ).order_by('name')

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Без движения'

    sheet.append([
        'Наименование',
        'Категория',
        'Поставщик',
        'Пол',
        'Размер',
        'Цвет',
        'Артикул',
        'Цена',
        'Количество',
        'Минимальный остаток',
        f'Нет продаж за {period_days_int} дней'
    ])

    for product in products:
        sheet.append([
            product.name,
            str(product.category),
            str(product.supplier) if product.supplier else '',
            product.gender,
            product.size,
            product.color,
            product.article,
            float(product.price),
            product.quantity,
            product.minimum_quantity,
            'Да'
        ])

    _style_header(sheet)
    _autosize_columns(sheet)

    return _make_excel_response(workbook, 'no_movement_products.xlsx')


@login_required
def action_log_list(request):
    if not (_is_admin(request.user) or _is_merchandiser(request.user)):
        return _forbidden(request)

    logs = ActionLog.objects.select_related('user', 'product').all().order_by('-created_at')[:200]

    return render(request, 'inventory/action_log_list.html', {'logs': logs})


@login_required
def create_sale(request):
    if not (_is_admin(request.user) or _is_seller(request.user)):
        return _forbidden(request)

    if request.method == 'POST':
        form = SaleForm(request.POST)
        if form.is_valid():
            sale = form.save(commit=False)
            product = sale.product

            if sale.quantity > product.quantity:
                messages.error(request, 'Недостаточно товара на складе для оформления продажи.')
            else:
                product.quantity -= sale.quantity
                product.save()
                sale.save()

                _log_action(
                    user=request.user,
                    action_type="CREATE_SALE",
                    product=product,
                    description=f'Оформлена продажа товара "{product.name}" в количестве {sale.quantity}'
                )

                messages.success(request, 'Продажа успешно оформлена.')
                return redirect('create_sale')
    else:
        form = SaleForm()

    return render(request, 'inventory/create_sale.html', {'form': form})


@login_required
def create_receipt(request):
    if not (_is_admin(request.user) or _is_merchandiser(request.user)):
        return _forbidden(request)

    if request.method == 'POST':
        form = StockReceiptForm(request.POST)
        if form.is_valid():
            receipt = form.save(commit=False)
            product = receipt.product
            product.quantity += receipt.quantity
            product.save()
            receipt.save()

            _log_action(
                user=request.user,
                action_type="CREATE_RECEIPT",
                product=product,
                description=f'Оформлено поступление товара "{product.name}" в количестве {receipt.quantity}'
            )

            messages.success(request, 'Поступление успешно сохранено.')
            return redirect('create_receipt')
    else:
        form = StockReceiptForm()

    return render(request, 'inventory/create_receipt.html', {'form': form})


@login_required
def product_history(request, product_id):
    if not (_is_admin(request.user) or _is_merchandiser(request.user) or _is_seller(request.user)):
        return _forbidden(request)

    product = get_object_or_404(Product, id=product_id)
    receipts = StockReceipt.objects.filter(product=product).order_by('-receipt_date', '-id')
    sales = Sale.objects.filter(product=product).order_by('-sale_date', '-id')

    context = {
        'product': product,
        'receipts': receipts,
        'sales': sales,
    }
    return render(request, 'inventory/product_history.html', context)


@login_required
def sales_report(request):
    if not (_is_admin(request.user) or _is_merchandiser(request.user)):
        return _forbidden(request)

    sales = Sale.objects.select_related('product').all()

    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()

    if start_date:
        sales = sales.filter(sale_date__gte=start_date)

    if end_date:
        sales = sales.filter(sale_date__lte=end_date)

    report = (
        sales.values('product__name')
        .annotate(total_sold=Coalesce(Sum('quantity'), 0))
        .order_by('-total_sold', 'product__name')
    )

    sales_chart_labels = [item['product__name'] for item in report]
    sales_chart_data = [item['total_sold'] for item in report]

    context = {
        'report': report,
        'start_date': start_date,
        'end_date': end_date,
        'sales_chart_labels': sales_chart_labels,
        'sales_chart_data': sales_chart_data,
    }
    return render(request, 'inventory/sales_report.html', context)


@login_required
def export_sales_report_excel(request):
    if not (_is_admin(request.user) or _is_merchandiser(request.user)):
        return _forbidden(request)

    sales = Sale.objects.select_related('product').all()

    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()

    if start_date:
        sales = sales.filter(sale_date__gte=start_date)

    if end_date:
        sales = sales.filter(sale_date__lte=end_date)

    report = (
        sales.values('product__name')
        .annotate(total_sold=Coalesce(Sum('quantity'), 0))
        .order_by('-total_sold', 'product__name')
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Продажи'

    sheet.append([
        'Товар',
        'Количество проданных единиц'
    ])

    for item in report:
        sheet.append([
            item['product__name'],
            item['total_sold']
        ])

    _style_header(sheet)
    _autosize_columns(sheet)

    return _make_excel_response(workbook, 'sales_report.xlsx')
from django.contrib.auth.models import User

def create_test_user(request):
    if not User.objects.filter(username='admin').exists():
        User.objects.create_superuser('admin', 'admin@mail.com', 'admin12345')
    return HttpResponse("User created")
